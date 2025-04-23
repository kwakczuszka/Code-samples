import pytest
import time
import json
import openpyxl
from selenium                                       import webdriver
from selenium.webdriver.common.by                   import By
from selenium.webdriver.common.action_chains        import ActionChains
from selenium.webdriver.support                     import expected_conditions as EC
from selenium.webdriver.support.wait                import WebDriverWait
from selenium.webdriver.common.keys                 import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities

class AutoZowie():
  def setup_method(self):
    config                          = json.load(open("autozowie.conf", "r"))

    self.URL                        = config["web"]["url"]
    self.passwd                     = config["web"]["passwd"]
    self.email                      = config["web"]["email"]

    self.XPATH_USE_SSO              = config["locators"]["xpath_use_sso"]
    self.XPATH_SSO_LOGIN            = config["locators"]["xpath_sso_login"]
    self.XPATH_EMAIL                = config["locators"]["xpath_email"]
    self.XPATH_PASSWD               = config["locators"]["xpath_passwd"]
    self.ID_SUBMIT_EMAIL            = config["locators"]["id_submit_email"]
    self.XPATH_SENDER_LIST          = config["locators"]["xpath_sender_list"]
    self.XPATH_NEW_CONVERSATION     = config["locators"]["xpath_new_conversation"]
    self.XPATH_SENDER_LIST_ELEMENT  = config["locators"]["xpath_sender_list_element"]
    self.XPATH_MAIL_RECIPIENT       = config["locators"]["xpath_mail_recipient"]
    self.XPATH_MAIL_SUBJECT         = config["locators"]["xpath_mail_subject"]
    self.XPATH_ATTACHMENT           = config["locators"]["xpath_attachment"]
    self.XPATH_SEND_BUTTON          = config["locators"]["xpath_send_button"]
    self.XPATH_SEND_NOW_BUTTON      = config["locators"]["xpath_send_now_button"]

    excel_path                      = config["files"]["excel_path"]
    workbook                        = openpyxl.load_workbook(excel_path)
    sheet                           = workbook.active 

    self.mails                      = [] 

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True):
      self.mails.append(row)  
    workbook.close()

    self.driver                     = webdriver.Chrome()

  def teardown_method(self):
    self.driver.quit()
  
  def auto_zowie(self):
    self.driver.implicitly_wait(120)                                                # <---- IN CASE YOU NEED MORE TIME INCREASE THIS NUMBER
    self.driver.get(self.URL)
    self.driver.set_window_size(1552, 936)
    self.driver.find_element(By.XPATH, self.XPATH_USE_SSO).click()  
    self.driver.find_element(By.NAME, "email").send_keys(self.email)
    self.driver.find_element(By.XPATH, self.XPATH_SSO_LOGIN).click()  
    self.driver.find_element(By.XPATH, self.XPATH_EMAIL).send_keys(self.email)
    self.driver.find_element(By.ID, self.ID_SUBMIT_EMAIL).click()
    self.driver.find_element(By.XPATH, self.XPATH_PASSWD).send_keys(self.passwd)
  
    for mail in self.mails:
      self.driver.find_element(By.XPATH, self.XPATH_NEW_CONVERSATION).click()       # START NEW CONVERSATION
      self.driver.implicitly_wait(5)
      time.sleep(1)

      self.driver.find_element(By.XPATH, self.XPATH_SENDER_LIST).click()            # OPEN SENDER SELECTION LIST
      time.sleep(2)              
      self.driver.find_element(By.XPATH, self.XPATH_SENDER_LIST_ELEMENT).click()   
      self.driver.find_element(By.XPATH, self.XPATH_MAIL_RECIPIENT).send_keys(mail[0]) 
      time.sleep(1)
      self.driver.find_element(By.XPATH, self.XPATH_MAIL_SUBJECT).send_keys("GR for PO "+str(mail[1]))
      iframe = self.driver.find_element(By.TAG_NAME, "iframe")
      self.driver.switch_to.frame(iframe)
      editable_body = self.driver.find_element(By.ID, "tinymce")
      editable_body.clear()
      message = "Hello\nPlease make a GR for attached invoice " + str(mail[2]) + " - thank you. \nBest regards,\nPTP PO Team\n\n[Please be advised, this message was sent automatically by system still in development]"
      editable_body.send_keys(message)
      self.driver.switch_to.default_content()
      self.driver.find_element(By.XPATH, self.XPATH_ATTACHMENT).send_keys(mail[3])
      time.sleep(5)                                                                 # WAIT FOR FILE TO UPLOAD

      self.driver.find_element(By.XPATH, self.XPATH_SEND_BUTTON).click()
      wait = WebDriverWait(self.driver, 10) 
      wait.until(EC.presence_of_element_located((By.XPATH, self.XPATH_SEND_NOW_BUTTON)))
      self.driver.find_element(By.XPATH, self.XPATH_SEND_NOW_BUTTON).click()

obj = AutoZowie()
obj.setup_method()
obj.auto_zowie()
obj.teardown_method()